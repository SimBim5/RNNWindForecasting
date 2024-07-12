import openpyxl
import os
from openpyxl.styles import PatternFill, Font, Alignment


def adjust_column_width(ws):
    """
    Adjust the column width based on the longest item in each column.
    """
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                # Consider the length of the cell value when adjusting column width
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adding a little extra space
        ws.column_dimensions[col[0].column_letter].width = adjusted_width


def apply_formatting(target_dir):
    for filename in os.listdir(target_dir):
        print(filename)
        if filename.endswith('.xlsx'):
            filepath = os.path.join(target_dir, filename)
            wb = openpyxl.load_workbook(filepath)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                
                # Set font size 12, Calibri, and center alignment for all cells
                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = Font(name='Calibri', size=12)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Format first row with blue background and light gray font, bold
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.fill = PatternFill(start_color="0065BD", end_color="0065BD", fill_type="solid")
                        cell.font = Font(name='Calibri', size=12, bold=True, color="EEECE1")
            
                adjust_column_width(ws)

            # Save the modified workbook
            wb.save(filepath)
    print("Formatting applied to all Excel files in the directory.")

# Example usage
target_directory = 'excel'
apply_formatting(target_directory)