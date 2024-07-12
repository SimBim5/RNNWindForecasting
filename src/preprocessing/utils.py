import numpy as np
import pandas as pd
import torch
import os
import matplotlib.pyplot as plt

from torch.utils.data import TensorDataset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows

def fill_data(data):
    return data.interpolate(method='linear', limit_area='inside')

def min_max_dataframe(data):
    return pd.DataFrame([data.max()])

def save_wind_as_tensors(input_sequences, ground_truth):
    input_sequences_tensor = torch.tensor(input_sequences, dtype=torch.float32)
    ground_truth_tensor = torch.tensor(ground_truth, dtype=torch.float32)
    return input_sequences_tensor, ground_truth_tensor

def save_weather_as_tensor(weather_data):
    weather_data_tensor = torch.tensor(weather_data, dtype=torch.float32)
    return weather_data_tensor

def save_as_dataset(input, ground_truth):
    dataset = TensorDataset(input, ground_truth)
    return dataset

def plot_correlation_matrix(corr_matrix, farm_names, input_sequences, save_index):
    import seaborn as sns
    sns.set_theme(style="dark")
    fig, axes = plt.subplots(1, 2, figsize=(18, 7))  # Use subplots with 1 row and 2 columns

    # Plotting the correlation matrix heatmap
    mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=1)
    sns.heatmap(corr_matrix, mask=mask, cmap="crest", vmax=1, vmin=-1,
                square=True, annot=True, fmt=".2f", linewidths=1, 
                xticklabels=farm_names, yticklabels=farm_names, 
                cbar_kws={"shrink": .5}, ax=axes[0])
        
    # Plotting the time series data
    time_points = np.arange(input_sequences.shape[0])
    series_cmap = sns.color_palette("cubehelix", n_colors=input_sequences.shape[1])
    for i, color in enumerate(series_cmap):
        axes[1].plot(time_points, input_sequences[:, i], label=f'{farm_names[i]}', color=color, linewidth=2)
        
    axes[1].legend(loc='upper right', fontsize='small')
    axes[1].grid(True)

    # Adjusting colors of ticklabels for consistency
    for ticklabel, color in zip(axes[0].get_yticklabels(), series_cmap):
        ticklabel.set_color(color)
    for ticklabel, color in zip(axes[0].get_xticklabels(), series_cmap):
        ticklabel.set_color(color)

    save_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'results', 'spatial')
    
    plt.tight_layout()
    plt.savefig(f'{save_dir}/{save_index+1}_correlation')
    plt.close()
    
def plot_distance_matrix(distance_metrics, farm_names, coordinates, save_index):
    import seaborn as sns
    sns.set_theme(style="dark")
    distance_matrix = np.zeros((len(farm_names), len(farm_names)))
    index = 0
    for i in range(len(farm_names)):
        for j in range(i+1, len(farm_names)):
            distance = distance_metrics[index]
            distance_matrix[i, j] = distance_matrix[j, i] = distance
            index += 1
  
    # Setting up the figure and axes for the subplot
    fig, axes = plt.subplots(1, 2, figsize=(18, 7))

    # Plotting the distance matrix heatmap
    mask = np.triu(np.ones_like(distance_matrix, dtype=bool), k=1)
    sns.heatmap(distance_matrix, mask=mask, cmap="crest",
                square=True, annot=True, fmt=".2f", linewidths=1, xticklabels=farm_names, 
                yticklabels=farm_names, cbar_kws={"shrink": .5}, ax=axes[0])
    
    latitudes = coordinates[:, 0]
    longitudes = coordinates[:, 1]

    # Plotting the scatter plot for farm locations
    color_values = np.linspace(0, 1, len(farm_names))
    scatter = axes[1].scatter(longitudes, latitudes, c=color_values, cmap='crest', s=100)
    for i, txt in enumerate(farm_names):
        axes[1].annotate(txt, (longitudes[i], latitudes[i]))
    axes[1].grid(True)
    
    series_cmap = sns.color_palette("cubehelix", n_colors=farm_names.shape[0])
    for ticklabel, color in zip(axes[0].get_yticklabels(), series_cmap):
        ticklabel.set_color(color)
    for ticklabel, color in zip(axes[0].get_xticklabels(), series_cmap):
        ticklabel.set_color(color)

    save_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'results', 'spatial')
    plt.tight_layout()
    plt.savefig(f'{save_dir}/{save_index+1}_distance')
    plt.close()
    
def save_lagged_correlations_excel(base_save_dir, example_index, lagged_correlations, max_lag, best_lags):
    save_dir = os.path.join(base_save_dir, f'{example_index+1}_Lagged_Correlations.xlsx')
    wb = Workbook()
    
    # Sheet 1: Correlations
    ws1 = wb.active
    ws1.title = "correlations"
    data_to_save = []
    for farm_pair, correlations in lagged_correlations.items():
        for lag, correlation in enumerate(correlations):
            data_to_save.append({
                'Farm Pair': farm_pair,
                'Lag': lag,
                'Correlation': correlation
            })
    df_to_save = pd.DataFrame(data_to_save)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_to_save, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(size=13)
            cell.alignment = Alignment(horizontal='center')
            if r_idx == 1:
                cell.fill = PatternFill(start_color='005293', end_color='005293', fill_type='solid')
                cell.font = Font(color='ffffff', size=13)

    for row in ws1.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    for column in ws1.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]
        if column:
            max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = max_length + 10
        ws1.column_dimensions[column[0].column_letter].width = adjusted_width
    
    start_row = 2
    row_increment = max_lag + 1

    for i in range(start_row, ws1.max_row, row_increment):
        end_row = i + max_lag
        range_string = f'C{i}:C{end_row}'
        if i % 2 == 0:
            ws1.conditional_formatting.add(range_string, ColorScaleRule(start_type='min', start_color='e37222', end_type='max', end_color='FFBEC0'))
        else:
            ws1.conditional_formatting.add(range_string,ColorScaleRule(start_type='min', start_color='ACAC9A', end_type='max', end_color='FBFAEA'))

    
    # Sheet 2: Best Lag
    ws2 = wb.create_sheet("best lag")
    best_lag_data = []
    best_lag_data = [dict(farm_pair=farm_pair, best_lag=lag, correlation=correlation) 
                        for farm_pair, (lag, correlation) in best_lags.items()]
    df_best_lag = pd.DataFrame(best_lag_data)
    
    for row in dataframe_to_rows(df_best_lag, index=False, header=True):
        ws2.append(row)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_best_lag, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(size=13)
            cell.alignment = Alignment(horizontal='center')
            if r_idx == 1:
                cell.fill = PatternFill(start_color='005293', end_color='005293', fill_type='solid')
                cell.font = Font(color='ffffff', size=13)

    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    for column in ws2.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]
        if column:
            max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = max_length + 10
        ws2.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(save_dir)
    
def adjust_std(arr, min_std=0.0001):
    if arr.std() < min_std:
        arr += np.random.normal(0, min_std, arr.shape)
    return arr

def safe_corrcoef(input_sequences):
    adjusted_sequences = np.apply_along_axis(adjust_std, axis=0, arr=input_sequences)
    return np.corrcoef(adjusted_sequences, rowvar=False)

def max_wind_data(data_path):
    df = pd.read_parquet(data_path)
    df = df.drop(columns=['SETTLEMENTDATE'])
    max_values = df.max()
    return max_values