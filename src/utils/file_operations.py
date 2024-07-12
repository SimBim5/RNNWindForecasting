from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
import os
import pandas as pd
import re
import logging

def get_cache_path(data_path, sequence_length, prediction_length, num_series, norm, mode, cluster_type):
    cache_dir = os.path.join('data', 'cache', 'wind_farm_data')
    os.makedirs(cache_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(data_path))[0]  
    if mode == "test_all":
        mode = "test"
    config_id = f"{base_name}_seq{sequence_length}_pred{prediction_length}_num{num_series}_norm{norm}_mode{mode}_cluster{cluster_type}"
    cache_file_name = f"{config_id}.pt"
    cache_path = os.path.join(cache_dir, cache_file_name)
    return cache_path

def get_weather_cache_path(data_path, sequence_length, prediction_length, num_series, norm, mode, cluster_type, config):
    cache_dir = os.path.join('data', 'cache', 'weather')
    os.makedirs(cache_dir, exist_ok=True)

    features = [
        "temp2m" if config["weather"]["temperature_2m"] else "",
        "rh2m" if config["weather"]["relative_humidity_2m"] else "",
        "rain" if config["weather"]["rain"] else "",
        "snow" if config["weather"]["snowfall"] else "",
        "pressure" if config["weather"]["pressure_msl"] else "",
        "surface_pressure" if config["weather"]["surface_pressure"] else "",
        "cloud_cover" if config["weather"]["cloud_cover"] else "",
        "cloud_cover_low" if config["weather"]["cloud_cover_low"] else "",
        "cloud_cover_mid" if config["weather"]["cloud_cover_mid"] else "",
        "cloud_cover_high" if config["weather"]["cloud_cover_high"] else "",
        "wind10m" if config["weather"]["wind_speed_10m"] else "",
        "wind100m" if config["weather"]["wind_speed_100m"] else "",
        "winddir10m" if config["weather"]["wind_direction_10m"] else "",
        "winddir100m" if config["weather"]["wind_direction_100m"] else "",
        "windgust10m" if config["weather"]["wind_gusts_10m"] else "",
        "soilmoisture0-7cm" if config["weather"]["soil_moisture_0_to_7cm"] else "",
        "soilmoisture7-28cm" if config["weather"]["soil_moisture_7_to_28cm"] else "",
        "soilmoisture28-100cm" if config["weather"]["soil_moisture_28_to_100cm"] else "",
        "soilmoisture100-255cm" if config["weather"]["soil_moisture_100_to_255cm"] else ""
    ]
    feature_str = "_".join(filter(None, features))    
    base_name = os.path.splitext(os.path.basename(data_path))[0]  
    if mode == "test_all":
        mode = "test"
    config_id_parts = [
        base_name, f"seq{sequence_length}", f"pred{prediction_length}", 
        f"num{num_series}", f"norm{norm}", f"mode{mode}", f"cluster{cluster_type}"
    ]
    if feature_str:
        config_id_parts.append(feature_str)
    config_id = "_".join(config_id_parts)
    cache_file_name = f"{config_id}.pt"
    cache_path = os.path.join(cache_dir, cache_file_name)
    return cache_path

def get_spatial_cache_path(mode, sequence_length, prediction_length, num_series, cluster_type):
    cache_dir = os.path.join('data', 'cache', 'spatial')
    os.makedirs(cache_dir, exist_ok=True)
    if mode == "test_all":
        mode = "test"
        
    config_id = f"{mode}_seq{sequence_length}_pred{prediction_length}_num{num_series}_cluster{cluster_type}"
    cache_file_name = f"{config_id}.pt"
    cache_path = os.path.join(cache_dir, cache_file_name)
    return cache_path

def generate_parameters_name(config):
    common_config = config['common']
    train_config = config.get('train', {})
    spatial_config = config.get('spatial', {}) 

    sequence_length = common_config.get('sequence_length', 'NA')
    prediction_length = common_config.get('prediction_length', 'NA')
    num_series = common_config.get('num_series', 'NA')
    hidden_size = common_config.get('hidden_size', 'NA')
    num_layers = common_config.get('num_layers', 'NA')
    num_directions = common_config.get('num_directions', 'NA')
    normalization = common_config.get('normalization', 'NA')
    epochs = train_config.get('epochs', 'NA')
    cluster_type = common_config['cluster_type']

    learning_rate = train_config.get('learning_rate', 'NA')
    loss_function = train_config.get('loss_function', 'NA')

    spatial_use = spatial_config.get('spatial_use', 'NA')

    filename = f"params_seq{sequence_length}_pred{prediction_length}_series{num_series}_hidden{hidden_size}_layers{num_layers}_directions{num_directions}_norm{normalization}_epochs{epochs}_lr{learning_rate}_cluster{cluster_type}_{loss_function}_spatial{spatial_use}"
    return filename

def generate_test_filename_from_config(config):
    test_config = config['test']
    plot_type = test_config.get('plot_type', 'NA')
    parameters_path = test_config.get('parameters_path', 'NA')
    filename = f"{parameters_path}_plot{plot_type}"
    return filename

def save_to_file(data, file_path):
    with open(file_path, 'w') as f:
        f.write(str(data))
        
def update_excel(data_df, save_dir="model_results.xlsx", sheet_name="Model Metrics"):
    if os.path.exists(save_dir):
        wb = load_workbook(filename=save_dir)
        # If the specified sheet exists, select it, otherwise add it
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            start_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title=sheet_name)
            start_row = 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        start_row = 1

    for r_idx, row in enumerate(dataframe_to_rows(data_df, index=False, header=(start_row==1)), start=start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(size=13)
            cell.alignment = Alignment(horizontal='center')
            if r_idx == 1:  # Header styling, only apply if it's the first row being added
                cell.fill = PatternFill(start_color='005293', end_color='005293', fill_type='solid')
                cell.font = Font(color='ffffff', size=13)

    if start_row == 1:
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
            ws.column_dimensions[column[0].column_letter].width = max_length

    wb.save(filename=save_dir)
    logging.info(f"Metrics successfully appended to {save_dir}.")
    
def parse_params_from_path(parameters_path):
    params_list = parameters_path.split('/')[1].split('_')    
    params_list = [extract_number(s) for s in params_list]
        
    params = {
        'Sequence Length': params_list[1],
        'Prediction Length': params_list[2],
        'Input Series': params_list[3],
        'Hidden Units': params_list[4],
        'LSTM Layers': params_list[5],
        'Directions': params_list[6],
        'Epochs': params_list[10],
        'Learning Rate': params_list[11],
        'Cluster': params_list[12],
        'Spatial': True if 'spatialTrue' in parameters_path else False if 'spatialFalse' in parameters_path else None
    }
    return params

def extract_number(input_str):
    pattern = r'-?\d+\.?\d*'
    matches = re.findall(pattern, input_str)
    if matches:
        num_str = matches[0]
        return float(num_str) if '.' in num_str else int(num_str)
    else:
        return input_str

def prepare_data_for_excel(parameters_path, metrics):
    params = parse_params_from_path(parameters_path)
    data = {**params, **metrics}
    return pd.DataFrame([data])

def extract_config_from_pathname(pathname):
    pattern = r'params_seq(\d+)_pred(\d+)_series(\d+)_hidden(\d+)_layers(\d+)_directions(\d+)_norm(\w+)_minmax_(\w+)_epochs(\d+)_lr([\d.]+)_cluster(nearest|random|capacity)_mean_squared_error_spatial(True|False)'
    match = re.search(pattern, pathname)
    
    if not match:
        raise ValueError("The pathname does not match the expected pattern.")
    
    data_folder = "data/wind_farm_data/"
    sequence_length = int(match.group(1))
    prediction_length = int(match.group(2))
    num_series = int(match.group(3))
    hidden_size = int(match.group(4))
    num_layers = int(match.group(5))
    num_directions = int(match.group(6))
    normalization = match.group(7) + "_minmax_" + match.group(8)
    epochs = int(match.group(9))
    learning_rate = float(match.group(10))
    cluster_type = match.group(11)
    spatial = match.group(12) == 'True'
   
    return data_folder, sequence_length, prediction_length, num_series, hidden_size, num_layers, num_directions, normalization, epochs, learning_rate, cluster_type, spatial
